
import openpyxl

path = "D:\Projects\pythonBasics\TestData.xlsx"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=1,column=1) # get first row and first column content
print(cell_obj.value)

# Print the total no of rows

print(sheet_obj.max_row)
print(sheet_obj.max_column)

# Loop will print all columns name

for i in range(1, sheet_obj.max_column+1):
    cell_obj = sheet_obj.cell(row=1,column=i)
    print(cell_obj.value)

